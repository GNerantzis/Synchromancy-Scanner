import sys
import traceback
import sqlite3
from datetime import datetime
from pathlib import Path
import ccxt
import requests
import pandas as pd
import numpy as np
import yfinance as yf

try:
    from tradingview_screener import Query
    HAS_TV_SCREENER = True
except Exception:
    Query = None
    HAS_TV_SCREENER = False

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QSpinBox, QLabel,
    QProgressBar, QHeaderView, QTabWidget, QComboBox, QCheckBox
)
from PySide6.QtCore import Qt, QRunnable, QThreadPool, Signal, QObject, QUrl
from PySide6.QtGui import QDesktopServices


# =====================================================
# SETTINGS
# =====================================================

COMMISSION = 0.001

DAILY_PREFERRED_TRADES = 14
DAILY_MINIMUM_TRADES = 2

HIGHER_PREFERRED_TRADES = 12
HIGHER_MINIMUM_TRADES = 2

DAILY_ATRS = list(range(5, 21))
WEEKLY_FACTORS = np.round(np.arange(1.7, 3.01, 0.1), 1)

MIN_BARS_CRYPTO = 50
MIN_BARS_DAILY = 60
MIN_BARS_WEEKLY = 80
YF_PERIOD = "14y"

# Asset-count value is preserved per asset class.
# Crypto and Stocks use market-cap sorted rank from #1 to #N.
# Real Assets uses local universe order from #1 to #N.
ASSET_CLASS_SETTINGS = {
    "Crypto": {"count": 50, "absolute_max": 5000},
    "Stocks": {"count": 100, "absolute_max": 1000},
    "Real Assets": {"count": 164, "absolute_max": 164},
}

# Stability weighted optimizer, matched to Pine version
# Synchromancy Stability Weighted 280526 WORKING LIGHT
USE_STABILITY_SELECTION = True
STABILITY_PENALTY = 0.8
MIN_STABILITY_SCORE = 0.25


# =====================================================
# FACTOR CACHE, SQLITE
# =====================================================

CACHE_WARNING_DAYS = 30


def app_folder():
    """Use the .exe folder when frozen, otherwise the .py folder."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


FACTOR_CACHE_DB = app_folder() / "factor_cache.sqlite"


def init_factor_cache():
    conn = sqlite3.connect(str(FACTOR_CACHE_DB))
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS factor_cache (
            symbol TEXT NOT NULL,
            market_type TEXT NOT NULL,
            timeframe TEXT NOT NULL,
            selected_parameter REAL NOT NULL,
            roi_winner_param REAL,
            stab_winner_param REAL,
            roi_winner_bt REAL,
            stab_winner_bt REAL,
            changed_by_stability INTEGER,
            used_trades INTEGER,
            all_trades INTEGER,
            valid INTEGER,
            last_optimized_date TEXT NOT NULL,
            PRIMARY KEY (symbol, market_type, timeframe)
        )
    """)
    conn.commit()
    conn.close()


def get_cached_parameter(symbol, market_type, timeframe):
    conn = sqlite3.connect(str(FACTOR_CACHE_DB))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT *
        FROM factor_cache
        WHERE symbol = ?
          AND market_type = ?
          AND timeframe = ?
    """, (symbol, market_type, timeframe))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None


def save_cached_parameter(symbol, market_type, timeframe, winner):
    """
    Important:
    Daily cache stores selected ATR length.
    Weekly cache stores selected fixed factor.
    """
    if timeframe in ["1w", "1M"]:
        selected_parameter = float(winner.get("factor", winner.get("param", 3.0)))
    else:
        selected_parameter = float(winner.get("atr", winner.get("param", 20)))

    conn = sqlite3.connect(str(FACTOR_CACHE_DB))
    cur = conn.cursor()
    cur.execute("""
        INSERT OR REPLACE INTO factor_cache (
            symbol,
            market_type,
            timeframe,
            selected_parameter,
            roi_winner_param,
            stab_winner_param,
            roi_winner_bt,
            stab_winner_bt,
            changed_by_stability,
            used_trades,
            all_trades,
            valid,
            last_optimized_date
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        symbol,
        market_type,
        timeframe,
        selected_parameter,
        none_to_float(winner.get("roi_winner_param")),
        none_to_float(winner.get("stab_winner_param")),
        none_to_float(winner.get("roi_winner_bt")),
        none_to_float(winner.get("stab_winner_bt")),
        1 if winner.get("changed_by_stability") else 0,
        none_to_int(winner.get("used_trades")),
        none_to_int(winner.get("all_trades")),
        1 if winner.get("valid") else 0,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))
    conn.commit()
    conn.close()


def none_to_float(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return float(value)


def none_to_int(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return int(value)


def cache_age_text(cache_row):
    if not cache_row:
        return "NEW"

    try:
        last = datetime.strptime(cache_row["last_optimized_date"], "%Y-%m-%d %H:%M:%S")
        days = (datetime.now() - last).days
    except Exception:
        return "UNKNOWN"

    if days > CACHE_WARNING_DAYS:
        return f"OLD {days}d"
    return f"{days}d"


def run_cached_logic(df, timeframe, cache_row):
    selected_parameter = float(cache_row["selected_parameter"])

    if timeframe in ["1w", "1M"]:
        atr_len = 10
        factor = selected_parameter
        factor_series = pd.Series(factor, index=df.index)
        st = supertrend_tv(df, factor_series, atr_len)
    else:
        atr_len = int(round(selected_parameter))
        factor_series = daily_factor(df, atr_len)
        st = supertrend_tv(df, factor_series, atr_len)
        factor_clean = factor_series.dropna()
        factor = round(float(factor_clean.iloc[-1]), 2) if not factor_clean.empty else np.nan

    if st.empty:
        return None

    return {
        "param": selected_parameter,
        "atr": atr_len,
        "factor": factor,
        "factor_now": factor,
        "bt": cache_row.get("stab_winner_bt"),
        "used_trades": cache_row.get("used_trades") if cache_row.get("used_trades") is not None else "NA",
        "all_trades": cache_row.get("all_trades") if cache_row.get("all_trades") is not None else "NA",
        "valid": bool(cache_row.get("valid")),
        "st": st,
        "roi_winner_param": cache_row.get("roi_winner_param"),
        "stab_winner_param": cache_row.get("stab_winner_param"),
        "roi_winner_bt": cache_row.get("roi_winner_bt"),
        "stab_winner_bt": cache_row.get("stab_winner_bt"),
        "changed_by_stability": bool(cache_row.get("changed_by_stability")),
        "mode": "CACHE"
    }


def run_logic_with_cache(df, timeframe, symbol, market_type, force_reoptimize=False):
    cache_row = None if force_reoptimize else get_cached_parameter(symbol, market_type, timeframe)

    if cache_row is not None:
        winner = run_cached_logic(df, timeframe, cache_row)
        if winner is not None:
            return winner, cache_age_text(cache_row)

    winner = run_logic(df, timeframe)
    if winner is None:
        return None, "NA"

    save_cached_parameter(symbol, market_type, timeframe, winner)
    return winner, "REFRESHED" if force_reoptimize else "NEW"


# =====================================================
# COMMODITY / REIT / REAL ASSET UNIVERSE
# =====================================================

COMMODITY_REIT_UNIVERSE = [{'category': 'Precious Metals ETF', 'description': 'Gold ETF', 'name': 'SPDR Gold Shares', 'symbol': 'GLD'},
 {'category': 'Precious Metals ETF', 'description': 'Gold ETF', 'name': 'iShares Gold Trust', 'symbol': 'IAU'},
 {'category': 'Precious Metals ETF',
  'description': 'Physical gold ETF',
  'name': 'abrdn Physical Gold Shares',
  'symbol': 'SGOL'},
 {'category': 'Precious Metals ETF',
  'description': 'Physical gold ETF',
  'name': 'GraniteShares Gold Trust',
  'symbol': 'BAR'},
 {'category': 'Precious Metals ETF', 'description': 'Gold ETF', 'name': 'VanEck Merk Gold Trust', 'symbol': 'OUNZ'},
 {'category': 'Precious Metals ETF', 'description': 'Silver ETF', 'name': 'iShares Silver Trust', 'symbol': 'SLV'},
 {'category': 'Precious Metals ETF',
  'description': 'Physical silver ETF',
  'name': 'abrdn Physical Silver Shares',
  'symbol': 'SIVR'},
 {'category': 'Precious Metals ETF',
  'description': 'Physical silver trust',
  'name': 'Sprott Physical Silver Trust',
  'symbol': 'PSLV'},
 {'category': 'Precious Metals ETF',
  'description': 'Platinum ETF',
  'name': 'abrdn Physical Platinum Shares',
  'symbol': 'PPLT'},
 {'category': 'Precious Metals ETF',
  'description': 'Palladium ETF',
  'name': 'abrdn Physical Palladium Shares',
  'symbol': 'PALL'},
 {'category': 'Gold Miners', 'description': 'Gold miners ETF', 'name': 'VanEck Gold Miners ETF', 'symbol': 'GDX'},
 {'category': 'Gold Miners',
  'description': 'Junior gold miners ETF',
  'name': 'VanEck Junior Gold Miners ETF',
  'symbol': 'GDXJ'},
 {'category': 'Silver Miners',
  'description': 'Silver miners ETF',
  'name': 'Global X Silver Miners ETF',
  'symbol': 'SIL'},
 {'category': 'Silver Miners',
  'description': 'Junior silver miners ETF',
  'name': 'Amplify Junior Silver Miners ETF',
  'symbol': 'SILJ'},
 {'category': 'Gold Miner', 'description': 'Major gold mining company', 'name': 'Newmont', 'symbol': 'NEM'},
 {'category': 'Gold Miner', 'description': 'Gold mining company', 'name': 'Agnico Eagle Mines', 'symbol': 'AEM'},
 {'category': 'Gold Miner', 'description': 'Major gold mining company', 'name': 'Barrick Gold', 'symbol': 'GOLD'},
 {'category': 'Gold Miner', 'description': 'Gold mining company', 'name': 'Kinross Gold', 'symbol': 'KGC'},
 {'category': 'Gold Miner', 'description': 'Gold mining company', 'name': 'AngloGold Ashanti', 'symbol': 'AU'},
 {'category': 'Royalty/Streaming',
  'description': 'Precious metals streaming company',
  'name': 'Wheaton Precious Metals',
  'symbol': 'WPM'},
 {'category': 'Royalty/Streaming',
  'description': 'Gold royalty and streaming company',
  'name': 'Franco-Nevada',
  'symbol': 'FNV'},
 {'category': 'Royalty/Streaming',
  'description': 'Precious metals royalty company',
  'name': 'Royal Gold',
  'symbol': 'RGLD'},
 {'category': 'Royalty/Streaming',
  'description': 'Gold royalty company',
  'name': 'Osisko Gold Royalties',
  'symbol': 'OR'},
 {'category': 'Silver Miner',
  'description': 'Silver and gold mining company',
  'name': 'Pan American Silver',
  'symbol': 'PAAS'},
 {'category': 'Silver Miner', 'description': 'Silver mining company', 'name': 'Hecla Mining', 'symbol': 'HL'},
 {'category': 'Silver Miner', 'description': 'Silver and gold mining company', 'name': 'Coeur Mining', 'symbol': 'CDE'},
 {'category': 'Silver Miner', 'description': 'Silver mining company', 'name': 'First Majestic Silver', 'symbol': 'AG'},
 {'category': 'Precious Metals Miner',
  'description': 'Platinum, palladium and gold miner',
  'name': 'Sibanye Stillwater',
  'symbol': 'SBSW'},
 {'category': 'Energy ETF',
  'description': 'US energy sector ETF',
  'name': 'Energy Select Sector SPDR',
  'symbol': 'XLE'},
 {'category': 'Energy ETF', 'description': 'Broad US energy ETF', 'name': 'Vanguard Energy ETF', 'symbol': 'VDE'},
 {'category': 'Energy ETF', 'description': 'US energy ETF', 'name': 'iShares US Energy ETF', 'symbol': 'IYE'},
 {'category': 'Oil Services ETF',
  'description': 'Oil services companies ETF',
  'name': 'VanEck Oil Services ETF',
  'symbol': 'OIH'},
 {'category': 'Oil & Gas ETF',
  'description': 'Oil and gas exploration ETF',
  'name': 'SPDR Oil & Gas Exploration ETF',
  'symbol': 'XOP'},
 {'category': 'Crude Oil ETF',
  'description': 'WTI crude oil futures ETF',
  'name': 'United States Oil Fund',
  'symbol': 'USO'},
 {'category': 'Crude Oil ETF',
  'description': 'Brent crude oil futures ETF',
  'name': 'United States Brent Oil Fund',
  'symbol': 'BNO'},
 {'category': 'Natural Gas ETF',
  'description': 'Natural gas futures ETF',
  'name': 'United States Natural Gas Fund',
  'symbol': 'UNG'},
 {'category': 'Leveraged Natural Gas ETF',
  'description': 'Leveraged natural gas ETF',
  'name': 'ProShares Ultra Bloomberg Natural Gas',
  'symbol': 'BOIL'},
 {'category': 'Inverse Natural Gas ETF',
  'description': 'Inverse leveraged natural gas ETF',
  'name': 'ProShares UltraShort Bloomberg Natural Gas',
  'symbol': 'KOLD'},
 {'category': 'Energy Commodity ETF',
  'description': 'Gasoline futures ETF',
  'name': 'United States Gasoline Fund',
  'symbol': 'UGA'},
 {'category': 'Oil Major', 'description': 'Integrated oil and gas major', 'name': 'Exxon Mobil', 'symbol': 'XOM'},
 {'category': 'Oil Major', 'description': 'Integrated oil and gas major', 'name': 'Chevron', 'symbol': 'CVX'},
 {'category': 'Oil Major', 'description': 'Integrated oil and gas major', 'name': 'Shell', 'symbol': 'SHEL'},
 {'category': 'Oil Major', 'description': 'Integrated oil and gas major', 'name': 'BP', 'symbol': 'BP'},
 {'category': 'Oil Major', 'description': 'Integrated oil and gas major', 'name': 'TotalEnergies', 'symbol': 'TTE'},
 {'category': 'Oil Producer',
  'description': 'Oil and gas exploration producer',
  'name': 'ConocoPhillips',
  'symbol': 'COP'},
 {'category': 'Oil Producer', 'description': 'US oil and gas producer', 'name': 'EOG Resources', 'symbol': 'EOG'},
 {'category': 'Oil Producer', 'description': 'Oil and gas producer', 'name': 'Occidental Petroleum', 'symbol': 'OXY'},
 {'category': 'Oil Producer', 'description': 'Oil and gas producer', 'name': 'Devon Energy', 'symbol': 'DVN'},
 {'category': 'Oil Producer', 'description': 'Permian oil producer', 'name': 'Diamondback Energy', 'symbol': 'FANG'},
 {'category': 'Oil Producer', 'description': 'Oil and gas producer', 'name': 'Hess', 'symbol': 'HES'},
 {'category': 'Refiner', 'description': 'Oil refining company', 'name': 'Marathon Petroleum', 'symbol': 'MPC'},
 {'category': 'Refiner', 'description': 'Oil refining company', 'name': 'Valero Energy', 'symbol': 'VLO'},
 {'category': 'Refiner', 'description': 'Oil refining and midstream company', 'name': 'Phillips 66', 'symbol': 'PSX'},
 {'category': 'LNG', 'description': 'Liquefied natural gas exporter', 'name': 'Cheniere Energy', 'symbol': 'LNG'},
 {'category': 'Natural Gas Producer', 'description': 'Natural gas producer', 'name': 'EQT Corp', 'symbol': 'EQT'},
 {'category': 'Natural Gas Producer',
  'description': 'Natural gas and NGL producer',
  'name': 'Antero Resources',
  'symbol': 'AR'},
 {'category': 'Natural Gas Producer',
  'description': 'Natural gas producer',
  'name': 'Range Resources',
  'symbol': 'RRC'},
 {'category': 'Natural Gas Producer',
  'description': 'Oil and natural gas producer',
  'name': 'Coterra Energy',
  'symbol': 'CTRA'},
 {'category': 'Pipeline/MLP ETF', 'description': 'Midstream MLP ETF', 'name': 'Alerian MLP ETF', 'symbol': 'AMLP'},
 {'category': 'Pipeline/Midstream',
  'description': 'Energy pipeline and midstream company',
  'name': 'Energy Transfer',
  'symbol': 'ET'},
 {'category': 'Pipeline/Midstream',
  'description': 'Midstream energy partnership',
  'name': 'Enterprise Products Partners',
  'symbol': 'EPD'},
 {'category': 'Pipeline/Midstream',
  'description': 'Natural gas pipeline operator',
  'name': 'Kinder Morgan',
  'symbol': 'KMI'},
 {'category': 'Pipeline/Midstream',
  'description': 'Natural gas infrastructure company',
  'name': 'Williams Companies',
  'symbol': 'WMB'},
 {'category': 'Pipeline/Midstream',
  'description': 'Natural gas liquids infrastructure',
  'name': 'ONEOK',
  'symbol': 'OKE'},
 {'category': 'Uranium ETF',
  'description': 'Uranium and nuclear energy ETF',
  'name': 'Global X Uranium ETF',
  'symbol': 'URA'},
 {'category': 'Uranium ETF',
  'description': 'Uranium miners ETF',
  'name': 'Sprott Uranium Miners ETF',
  'symbol': 'URNM'},
 {'category': 'Uranium Trust',
  'description': 'Physical uranium trust',
  'name': 'Sprott Physical Uranium Trust',
  'symbol': 'SRUUF'},
 {'category': 'Uranium Miner', 'description': 'Major uranium producer', 'name': 'Cameco', 'symbol': 'CCJ'},
 {'category': 'Uranium Miner',
  'description': 'US uranium and rare earths company',
  'name': 'Energy Fuels',
  'symbol': 'UUUU'},
 {'category': 'Uranium Developer',
  'description': 'Uranium development company',
  'name': 'NexGen Energy',
  'symbol': 'NXE'},
 {'category': 'Uranium Developer',
  'description': 'Uranium exploration and development',
  'name': 'Denison Mines',
  'symbol': 'DNN'},
 {'category': 'Uranium Miner', 'description': 'US uranium company', 'name': 'Uranium Energy Corp', 'symbol': 'UEC'},
 {'category': 'Copper ETF',
  'description': 'Copper futures ETF',
  'name': 'United States Copper Index Fund',
  'symbol': 'CPER'},
 {'category': 'Copper Miners ETF',
  'description': 'Copper miners ETF',
  'name': 'Global X Copper Miners ETF',
  'symbol': 'COPX'},
 {'category': 'Base Metals ETF',
  'description': 'Base metals futures ETF',
  'name': 'Invesco DB Base Metals Fund',
  'symbol': 'DBB'},
 {'category': 'Nickel ETN', 'description': 'Nickel exposure ETN', 'name': 'iPath Nickel ETN', 'symbol': 'JJN'},
 {'category': 'Aluminum ETN', 'description': 'Aluminum exposure ETN', 'name': 'iPath Aluminum ETN', 'symbol': 'JJU'},
 {'category': 'Copper Miner',
  'description': 'Copper and gold mining company',
  'name': 'Freeport-McMoRan',
  'symbol': 'FCX'},
 {'category': 'Copper Miner', 'description': 'Copper mining company', 'name': 'Southern Copper', 'symbol': 'SCCO'},
 {'category': 'Diversified Miner',
  'description': 'Copper, zinc, steelmaking coal miner',
  'name': 'Teck Resources',
  'symbol': 'TECK'},
 {'category': 'Diversified Miner',
  'description': 'Global diversified mining company',
  'name': 'BHP Group',
  'symbol': 'BHP'},
 {'category': 'Diversified Miner',
  'description': 'Global diversified mining company',
  'name': 'Rio Tinto',
  'symbol': 'RIO'},
 {'category': 'Iron Ore Miner', 'description': 'Iron ore and base metals miner', 'name': 'Vale', 'symbol': 'VALE'},
 {'category': 'Steel ETF', 'description': 'Steel producers ETF', 'name': 'VanEck Steel ETF', 'symbol': 'SLX'},
 {'category': 'Steel Producer', 'description': 'Steel producer', 'name': 'Nucor', 'symbol': 'NUE'},
 {'category': 'Steel Producer', 'description': 'Steel producer', 'name': 'Steel Dynamics', 'symbol': 'STLD'},
 {'category': 'Steel Producer', 'description': 'Steel producer', 'name': 'US Steel', 'symbol': 'X'},
 {'category': 'Steel/Iron Ore',
  'description': 'Steel and iron ore company',
  'name': 'Cleveland-Cliffs',
  'symbol': 'CLF'},
 {'category': 'Lithium ETF',
  'description': 'Lithium and battery technology ETF',
  'name': 'Global X Lithium & Battery Tech ETF',
  'symbol': 'LIT'},
 {'category': 'Lithium Producer',
  'description': 'Lithium and specialty chemicals company',
  'name': 'Albemarle',
  'symbol': 'ALB'},
 {'category': 'Lithium Producer',
  'description': 'Lithium, fertilizers and chemicals',
  'name': 'Sociedad Quimica y Minera',
  'symbol': 'SQM'},
 {'category': 'Lithium Developer',
  'description': 'Lithium development company',
  'name': 'Piedmont Lithium',
  'symbol': 'PLL'},
 {'category': 'Rare Earth ETF',
  'description': 'Rare earth and strategic metals ETF',
  'name': 'VanEck Rare Earth ETF',
  'symbol': 'REMX'},
 {'category': 'Rare Earths', 'description': 'Rare earth materials company', 'name': 'MP Materials', 'symbol': 'MP'},
 {'category': 'Agriculture ETF',
  'description': 'Broad agriculture futures ETF',
  'name': 'Invesco DB Agriculture Fund',
  'symbol': 'DBA'},
 {'category': 'Agribusiness ETF',
  'description': 'Agriculture businesses ETF',
  'name': 'VanEck Agribusiness ETF',
  'symbol': 'MOO'},
 {'category': 'Grain ETF', 'description': 'Wheat futures ETF', 'name': 'Teucrium Wheat Fund', 'symbol': 'WEAT'},
 {'category': 'Grain ETF', 'description': 'Corn futures ETF', 'name': 'Teucrium Corn Fund', 'symbol': 'CORN'},
 {'category': 'Grain ETF', 'description': 'Soybean futures ETF', 'name': 'Teucrium Soybean Fund', 'symbol': 'SOYB'},
 {'category': 'Soft Commodity ETN', 'description': 'Coffee exposure ETN', 'name': 'iPath Coffee ETN', 'symbol': 'JO'},
 {'category': 'Soft Commodity ETF',
  'description': 'Sugar futures ETF',
  'name': 'Teucrium Sugar Fund',
  'symbol': 'CANE'},
 {'category': 'Soft Commodity ETN', 'description': 'Cocoa exposure ETN', 'name': 'iPath Cocoa ETN', 'symbol': 'NIB'},
 {'category': 'Soft Commodity ETN', 'description': 'Cotton exposure ETN', 'name': 'iPath Cotton ETN', 'symbol': 'BAL'},
 {'category': 'Fertilizer',
  'description': 'Phosphate and potash fertilizer producer',
  'name': 'Mosaic',
  'symbol': 'MOS'},
 {'category': 'Fertilizer', 'description': 'Fertilizer and agriculture retailer', 'name': 'Nutrien', 'symbol': 'NTR'},
 {'category': 'Fertilizer', 'description': 'Nitrogen fertilizer producer', 'name': 'CF Industries', 'symbol': 'CF'},
 {'category': 'Agribusiness',
  'description': 'Agricultural processing and trading',
  'name': 'Archer-Daniels-Midland',
  'symbol': 'ADM'},
 {'category': 'Agribusiness', 'description': 'Agribusiness and food company', 'name': 'Bunge', 'symbol': 'BG'},
 {'category': 'Agriculture Equipment',
  'description': 'Farm and construction equipment',
  'name': 'Deere',
  'symbol': 'DE'},
 {'category': 'Agriculture/Rural Retail',
  'description': 'Rural lifestyle retailer',
  'name': 'Tractor Supply',
  'symbol': 'TSCO'},
 {'category': 'Broad Commodity ETF',
  'description': 'Broad commodity futures ETF',
  'name': 'Invesco DB Commodity Index Tracking Fund',
  'symbol': 'DBC'},
 {'category': 'Broad Commodity ETF',
  'description': 'Broad commodity strategy ETF',
  'name': 'Direxion Auspice Broad Commodity Strategy',
  'symbol': 'COM'},
 {'category': 'Broad Commodity ETF',
  'description': 'Broad commodity strategy ETF',
  'name': 'Invesco Optimum Yield Diversified Commodity Strategy',
  'symbol': 'PDBC'},
 {'category': 'Broad Commodity ETF',
  'description': 'Broad commodity index ETF',
  'name': 'iShares S&P GSCI Commodity ETF',
  'symbol': 'GSG'},
 {'category': 'Broad Commodity ETF',
  'description': 'Broad commodity strategy ETF',
  'name': 'abrdn Bloomberg All Commodity Strategy',
  'symbol': 'BCI'},
 {'category': 'REIT ETF', 'description': 'Broad US REIT ETF', 'name': 'Vanguard Real Estate ETF', 'symbol': 'VNQ'},
 {'category': 'REIT ETF', 'description': 'Broad US REIT ETF', 'name': 'Schwab US REIT ETF', 'symbol': 'SCHH'},
 {'category': 'REIT ETF', 'description': 'US real estate ETF', 'name': 'iShares US Real Estate ETF', 'symbol': 'IYR'},
 {'category': 'REIT ETF',
  'description': 'S&P 500 real estate sector ETF',
  'name': 'Real Estate Select Sector SPDR',
  'symbol': 'XLRE'},
 {'category': 'REIT ETF', 'description': 'US REIT ETF', 'name': 'SPDR Dow Jones REIT ETF', 'symbol': 'RWR'},
 {'category': 'Triple Net REIT',
  'description': 'Monthly dividend retail net lease REIT',
  'name': 'Realty Income',
  'symbol': 'O'},
 {'category': 'Triple Net REIT', 'description': 'Retail net lease REIT', 'name': 'NNN REIT', 'symbol': 'NNN'},
 {'category': 'Triple Net REIT', 'description': 'Retail net lease REIT', 'name': 'Agree Realty', 'symbol': 'ADC'},
 {'category': 'Net Lease REIT', 'description': 'Diversified net lease REIT', 'name': 'W. P. Carey', 'symbol': 'WPC'},
 {'category': 'Industrial REIT', 'description': 'Logistics and warehouse REIT', 'name': 'Prologis', 'symbol': 'PLD'},
 {'category': 'Industrial REIT',
  'description': 'Industrial real estate REIT',
  'name': 'First Industrial Realty Trust',
  'symbol': 'FR'},
 {'category': 'Industrial REIT',
  'description': 'Industrial real estate REIT',
  'name': 'Terreno Realty',
  'symbol': 'TRNO'},
 {'category': 'Industrial REIT',
  'description': 'Southern California industrial REIT',
  'name': 'Rexford Industrial Realty',
  'symbol': 'REXR'},
 {'category': 'Data Center REIT', 'description': 'Global data center REIT', 'name': 'Equinix', 'symbol': 'EQIX'},
 {'category': 'Data Center REIT', 'description': 'Data center REIT', 'name': 'Digital Realty', 'symbol': 'DLR'},
 {'category': 'Tower REIT', 'description': 'Telecom tower REIT', 'name': 'American Tower', 'symbol': 'AMT'},
 {'category': 'Tower REIT', 'description': 'Telecom tower REIT', 'name': 'Crown Castle', 'symbol': 'CCI'},
 {'category': 'Tower REIT', 'description': 'Telecom tower REIT', 'name': 'SBA Communications', 'symbol': 'SBAC'},
 {'category': 'Residential REIT', 'description': 'Apartment REIT', 'name': 'AvalonBay Communities', 'symbol': 'AVB'},
 {'category': 'Residential REIT', 'description': 'Apartment REIT', 'name': 'Equity Residential', 'symbol': 'EQR'},
 {'category': 'Residential REIT',
  'description': 'Apartment REIT',
  'name': 'Mid-America Apartment Communities',
  'symbol': 'MAA'},
 {'category': 'Residential REIT', 'description': 'Apartment REIT', 'name': 'Essex Property Trust', 'symbol': 'ESS'},
 {'category': 'Healthcare REIT',
  'description': 'Senior housing and healthcare REIT',
  'name': 'Welltower',
  'symbol': 'WELL'},
 {'category': 'Healthcare REIT',
  'description': 'Healthcare and senior housing REIT',
  'name': 'Ventas',
  'symbol': 'VTR'},
 {'category': 'Healthcare REIT',
  'description': 'Healthcare real estate REIT',
  'name': 'Healthpeak Properties',
  'symbol': 'DOC'},
 {'category': 'Healthcare REIT',
  'description': 'Skilled nursing REIT',
  'name': 'Omega Healthcare Investors',
  'symbol': 'OHI'},
 {'category': 'Self Storage REIT', 'description': 'Self storage REIT', 'name': 'Public Storage', 'symbol': 'PSA'},
 {'category': 'Self Storage REIT', 'description': 'Self storage REIT', 'name': 'Extra Space Storage', 'symbol': 'EXR'},
 {'category': 'Self Storage REIT', 'description': 'Self storage REIT', 'name': 'CubeSmart', 'symbol': 'CUBE'},
 {'category': 'Timber REIT', 'description': 'Timberland REIT', 'name': 'Weyerhaeuser', 'symbol': 'WY'},
 {'category': 'Timber REIT', 'description': 'Timberland REIT', 'name': 'PotlatchDeltic', 'symbol': 'PCH'},
 {'category': 'Farmland REIT', 'description': 'Farmland REIT', 'name': 'Gladstone Land', 'symbol': 'LAND'},
 {'category': 'Farmland REIT', 'description': 'Farmland REIT', 'name': 'Farmland Partners', 'symbol': 'FPI'},
 {'category': 'Hotel REIT', 'description': 'Hotel and lodging REIT', 'name': 'Host Hotels & Resorts', 'symbol': 'HST'},
 {'category': 'Hotel REIT',
  'description': 'Hotel and entertainment REIT',
  'name': 'Ryman Hospitality Properties',
  'symbol': 'RHP'},
 {'category': 'Casino REIT',
  'description': 'Casino and entertainment property REIT',
  'name': 'VICI Properties',
  'symbol': 'VICI'},
 {'category': 'Casino REIT',
  'description': 'Casino property REIT',
  'name': 'Gaming and Leisure Properties',
  'symbol': 'GLPI'},
 {'category': 'Mortgage REIT', 'description': 'Mortgage REIT', 'name': 'Annaly Capital Management', 'symbol': 'NLY'},
 {'category': 'Mortgage REIT', 'description': 'Mortgage REIT', 'name': 'AGNC Investment', 'symbol': 'AGNC'},
 {'category': 'Infrastructure Fund',
  'description': 'Infrastructure closed-end fund',
  'name': 'Cohen & Steers Infrastructure Fund',
  'symbol': 'UTF'},
 {'category': 'Infrastructure',
  'description': 'Global infrastructure assets',
  'name': 'Brookfield Infrastructure Partners',
  'symbol': 'BIP'},
 {'category': 'Infrastructure',
  'description': 'Global infrastructure assets',
  'name': 'Brookfield Infrastructure Corporation',
  'symbol': 'BIPC'},
 {'category': 'Utility',
  'description': 'Electric utility and renewable energy',
  'name': 'NextEra Energy',
  'symbol': 'NEE'},
 {'category': 'Utility', 'description': 'Electric and gas utility', 'name': 'Duke Energy', 'symbol': 'DUK'},
 {'category': 'Utility', 'description': 'Electric and gas utility', 'name': 'Southern Company', 'symbol': 'SO'},
 {'category': 'Utility', 'description': 'Electric utility', 'name': 'American Electric Power', 'symbol': 'AEP'},
 {'category': 'Utility', 'description': 'Electric and gas utility', 'name': 'Dominion Energy', 'symbol': 'D'},
 {'category': 'Utility', 'description': 'Electric utility', 'name': 'Exelon', 'symbol': 'EXC'},
 {'category': 'Utility', 'description': 'Energy infrastructure and utility', 'name': 'Sempra', 'symbol': 'SRE'}]


# =====================================================
# BYBIT SPOT, CRYPTO
# =====================================================

bybit_exchange = ccxt.bybit({
    "enableRateLimit": True,
    "options": {"defaultType": "spot"}
})

MARKETS_LOADED = False


def ensure_markets_loaded():
    global MARKETS_LOADED
    if not MARKETS_LOADED:
        bybit_exchange.load_markets()
        MARKETS_LOADED = True


def get_crypto_route(base_symbol):
    base_symbol = str(base_symbol).upper()

    aliases = {
        "WETH": "ETH",
        "WBTC": "BTC",
        "BTT": "BTTC"
    }

    base_symbol = aliases.get(base_symbol, base_symbol)
    pair = f"{base_symbol}/USDT"

    return {
        "exchange": bybit_exchange,
        "pair": pair,
        "tv_symbol": f"BYBIT:{base_symbol}USDT",
        "display": pair
    }


def get_coins_page(page):
    try:
        url = "https://api.coingecko.com/api/v3/coins/markets"
        params = {
            "vs_currency": "usd",
            "order": "market_cap_desc",
            "per_page": 250,
            "page": page
        }

        r = requests.get(url, params=params, timeout=15)

        if r.status_code != 200:
            return []

        return r.json()

    except Exception:
        return []


def get_crypto_universe(min_rank, max_rank):
    coins = []
    page = 1

    # Important: max_rank is candidate count, not final valid result count.
    # We over-fetch CoinGecko candidates because many top coins are stablecoins,
    # have no Bybit spot pair, have too little OHLCV history, or fail trade-count rules.
    while len(coins) < max_rank:
        data = get_coins_page(page)
        if not data:
            break
        coins.extend(data)
        page += 1

    seen = set()
    unique = []

    for coin in coins:
        cid = coin.get("id")
        if cid and cid not in seen:
            seen.add(cid)
            unique.append(coin)

    return unique[min_rank - 1:max_rank]


def fetch_crypto(route, timeframe, limit=1000):
    try:
        ensure_markets_loaded()

        exchange = route["exchange"]
        pair = route["pair"]

        if pair not in exchange.markets:
            return None

        market = exchange.markets[pair]

        if not market.get("spot"):
            return None

        if market.get("swap") or market.get("future"):
            return None

        if timeframe in ["1w", "1M"]:
            data_start = "2011-01-01"
        else:
            data_start = "2022-01-01"

        since = exchange.parse8601(f"{data_start}T00:00:00Z")
        all_bars = []

        while True:
            bars = exchange.fetch_ohlcv(pair, timeframe=timeframe, since=since, limit=limit)

            if not bars:
                break

            all_bars.extend(bars)
            last_ts = bars[-1][0]

            if last_ts <= since:
                break

            since = last_ts + 1

            if len(bars) < limit:
                break

            if len(all_bars) > 10000:
                break

        if not all_bars:
            return None

        df = pd.DataFrame(all_bars, columns=["time", "open", "high", "low", "close", "volume"])
        df = df.drop_duplicates(subset=["time"])
        df["time"] = pd.to_datetime(df["time"], unit="ms")
        df = df.dropna().reset_index(drop=True)
        return df

    except Exception:
        traceback.print_exc()
        return None


# =====================================================
# STOCKS / REAL ASSETS, YAHOO + TRADINGVIEW SCREENER
# =====================================================

def get_tv_stocks(max_rank):
    if not HAS_TV_SCREENER:
        return pd.DataFrame()

    try:
        _, df = (
            Query()
            .select("name", "description", "exchange", "market_cap_basic")
            .order_by("market_cap_basic", ascending=False)
            .limit(max_rank)
            .get_scanner_data()
        )
        return df
    except Exception:
        traceback.print_exc()
        return pd.DataFrame()


def get_stock_universe(min_rank, max_rank):
    df = get_tv_stocks(max_rank)
    if df.empty:
        return []
    return df.iloc[min_rank - 1:max_rank].to_dict("records")


def get_real_asset_universe(min_rank, max_rank):
    return COMMODITY_REIT_UNIVERSE[min_rank - 1:max_rank]


def clean_yfinance_symbol(symbol):
    symbol = str(symbol).split(":")[-1].strip().upper()
    return symbol.replace(".", "-")


def clean_tv_symbol(symbol, exchange=None):
    symbol = str(symbol).split(":")[-1].strip().upper()
    if exchange:
        return f"{str(exchange).strip()}:{symbol}"
    return symbol


def fetch_yfinance(symbol):
    try:
        df = yf.download(
            symbol,
            period=YF_PERIOD,
            interval="1d",
            auto_adjust=False,
            progress=False,
            threads=False
        )

        if df is None or len(df) == 0:
            return None

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)

        df = df.rename(columns={
            "Open": "open",
            "High": "high",
            "Low": "low",
            "Close": "close",
            "Volume": "volume"
        })

        df = df.reset_index()
        date_col = "Date" if "Date" in df.columns else "Datetime"
        df["time"] = pd.to_datetime(df[date_col])
        df = df[["time", "open", "high", "low", "close", "volume"]]
        df = df.dropna().reset_index(drop=True)
        return df

    except Exception:
        return None


def to_weekly(df):
    df = df.copy().set_index("time")
    weekly = df.resample("W-FRI").agg({
        "open": "first",
        "high": "max",
        "low": "min",
        "close": "last",
        "volume": "sum"
    })
    return weekly.dropna().reset_index()


def drop_unconfirmed_signal_bar(df, timeframe):
    """
    For mechanical scans, ignore the currently forming signal candle.

    This matches the Pine visual freeze idea:
    - Daily scans use the last confirmed daily candle.
    - Weekly/monthly scans use the last confirmed higher-timeframe candle.

    This prevents the selected Daily ATR, Weekly factor, trend state and flip metrics
    from changing because of a still-forming candle.
    """
    if df is None or df.empty:
        return df

    if timeframe in ["1d", "1w", "1M"]:
        if len(df) <= 1:
            return df.iloc[0:0].copy()
        return df.iloc[:-1].copy().reset_index(drop=True)

    return df


# Backwards-compatible alias for old calls.
def drop_unconfirmed_higher_tf_bar(df, timeframe):
    return drop_unconfirmed_signal_bar(df, timeframe)


# =====================================================
# INDICATORS
# =====================================================

def rma_tv(series, length):
    values = series.to_numpy(dtype=float)
    out = np.full(len(values), np.nan)

    if len(values) < length:
        return pd.Series(out, index=series.index)

    first = np.nanmean(values[:length])
    out[length - 1] = first

    for i in range(length, len(values)):
        out[i] = (out[i - 1] * (length - 1) + values[i]) / length

    return pd.Series(out, index=series.index)


def atr(df, length):
    prev_close = df["close"].shift(1)

    tr = pd.concat([
        df["high"] - df["low"],
        (df["high"] - prev_close).abs(),
        (df["low"] - prev_close).abs()
    ], axis=1).max(axis=1)

    return rma_tv(tr, length)


def get_dynamic_factor(current_atr, prior_atr):
    if pd.isna(current_atr) or pd.isna(prior_atr) or prior_atr == 0:
        return 3.0

    zone_size = prior_atr / 4.0

    if current_atr <= zone_size * 1.0:
        return 5.0
    elif current_atr <= zone_size * 2.0:
        return 3.0
    elif current_atr <= zone_size * 3.0:
        return 1.5
    else:
        return 1.0


def daily_factor(df, atr_len):
    current_atr = atr(df, atr_len)
    prior_atr = current_atr.shift(1)
    raw = []

    for c, p in zip(current_atr, prior_atr):
        raw.append(get_dynamic_factor(c, p))

    raw = pd.Series(raw, index=df.index)
    return 6.0 - raw


def supertrend_tv(df, factor_series, atr_len):
    df = df.copy().reset_index(drop=True)
    atr_series = atr(df, atr_len)
    factor_series = pd.Series(factor_series).reset_index(drop=True)
    hl2 = (df["high"] + df["low"]) / 2

    upper_basic = hl2 + factor_series * atr_series
    lower_basic = hl2 - factor_series * atr_series

    upper_band = pd.Series(np.nan, index=df.index)
    lower_band = pd.Series(np.nan, index=df.index)
    direction = pd.Series(np.nan, index=df.index)
    supertrend_line = pd.Series(np.nan, index=df.index)

    for i in range(len(df)):
        if pd.isna(atr_series.iloc[i]):
            continue

        if i == 0 or pd.isna(supertrend_line.iloc[i - 1]):
            upper_band.iloc[i] = upper_basic.iloc[i]
            lower_band.iloc[i] = lower_basic.iloc[i]
            direction.iloc[i] = 1
            supertrend_line.iloc[i] = upper_band.iloc[i]
            continue

        prev_lower = lower_band.iloc[i - 1]
        prev_upper = upper_band.iloc[i - 1]

        if lower_basic.iloc[i] > prev_lower or df["close"].iloc[i - 1] < prev_lower:
            lower_band.iloc[i] = lower_basic.iloc[i]
        else:
            lower_band.iloc[i] = prev_lower

        if upper_basic.iloc[i] < prev_upper or df["close"].iloc[i - 1] > prev_upper:
            upper_band.iloc[i] = upper_basic.iloc[i]
        else:
            upper_band.iloc[i] = prev_upper

        prev_st = supertrend_line.iloc[i - 1]

        if prev_st == prev_upper:
            direction.iloc[i] = -1 if df["close"].iloc[i] > upper_band.iloc[i] else 1
        else:
            direction.iloc[i] = 1 if df["close"].iloc[i] < lower_band.iloc[i] else -1

        supertrend_line.iloc[i] = lower_band.iloc[i] if direction.iloc[i] == -1 else upper_band.iloc[i]

    df["direction"] = direction
    df["trend"] = np.where(direction < 0, 1, -1)
    df.loc[direction.isna(), "trend"] = np.nan
    df["supertrend"] = supertrend_line
    df = df.dropna(subset=["trend", "supertrend"]).reset_index(drop=True)
    df["trend"] = df["trend"].astype(int)
    return df


# =====================================================
# BACKTEST ENGINE
# =====================================================

def calculate_trade_count_return(df, preferred_trades, minimum_trades):
    entry = None
    in_trade = False
    returns = []

    for i in range(len(df)):
        bull = df["trend"].iloc[i] == 1

        if i == 0:
            bull_flip = False
            bear_flip = False
        else:
            bull_flip = df["trend"].iloc[i - 1] == -1 and df["trend"].iloc[i] == 1
            bear_flip = df["trend"].iloc[i - 1] == 1 and df["trend"].iloc[i] == -1

        if ((i == 0 and bull) or bull_flip) and not in_trade:
            entry = df["close"].iloc[i]
            in_trade = True

        if bear_flip and in_trade and entry is not None:
            r = (df["close"].iloc[i] / entry) * (1.0 - COMMISSION * 2.0) - 1.0
            returns.append(r)
            entry = None
            in_trade = False

    trade_count = len(returns)
    safe_preferred = max(preferred_trades, 1)
    safe_minimum = max(minimum_trades, 1)
    enough = trade_count >= safe_minimum
    used_trades = min(trade_count, safe_preferred) if enough else 0

    if not enough or used_trades <= 0:
        return None, used_trades, trade_count, False

    selected_returns = returns[-used_trades:]
    equity = 1.0

    for r in selected_returns:
        equity *= 1.0 + r

    bt_roi = (equity - 1.0) * 100.0
    return bt_roi, used_trades, trade_count, True


def safe_plateau_score(roi, neighbor_avg):
    if neighbor_avg is None or pd.isna(neighbor_avg):
        return 1.0

    if neighbor_avg <= 0:
        return 1.0

    score = roi / neighbor_avg

    if pd.isna(score) or np.isinf(score):
        return 1.0

    return max(float(score), MIN_STABILITY_SCORE)


def label_plateau(score):
    if score > 1.25:
        return "SPIKE"
    elif score > 1.10:
        return "PEAKED"
    return "PLATEAU"


def add_stability_metrics(results):
    """
    Match the Pine logic from Synchromancy Stability Weighted 280526.

    Important detail: stability compares the immediate neighbouring parameter slots.
    It does not skip over invalid settings to find the next valid setting.
    Daily: ATR 5 compares only ATR 6, ATR 6 compares ATR 5 and 7, etc.
    Weekly: factor 1.7 compares only 1.8, factor 1.8 compares 1.7 and 1.9, etc.
    """
    ordered = sorted(results, key=lambda x: x["param"])

    for i, r in enumerate(ordered):
        left = ordered[i - 1] if i > 0 else None
        right = ordered[i + 1] if i < len(ordered) - 1 else None

        neighbors = []

        if left is not None and left.get("valid") and left.get("bt") is not None:
            neighbors.append(left["bt"])

        if right is not None and right.get("valid") and right.get("bt") is not None:
            neighbors.append(right["bt"])

        neighbor_avg = float(np.mean(neighbors)) if neighbors else None
        plateau_score = safe_plateau_score(r.get("bt"), neighbor_avg)

        if r.get("bt") is None:
            stability_score = None
        else:
            stability_score = r["bt"] / (plateau_score ** STABILITY_PENALTY)

        r["neighbor_avg"] = neighbor_avg
        r["plateau_score"] = plateau_score
        r["stability_score"] = stability_score
        r["stability_label"] = label_plateau(plateau_score)

    return [x for x in ordered if x["valid"] and x["bt"] is not None]


def select_roi_winner(valid):
    if not valid:
        return None
    return max(valid, key=lambda x: x["bt"])


def select_stability_winner(valid):
    if not valid:
        return None
    return max(valid, key=lambda x: -999999.0 if x.get("stability_score") is None else x["stability_score"])


def default_no_bt_result(results):
    """
    If there is no valid BT result, do not let the optimizer accidentally pick
    the first tested parameter, like factor 1.7.

    Higher TF default: ATR 10, factor 3.0.
    Daily and below default: ATR 20, using the ATR 20 dynamic result if available.
    """
    if not results:
        return None

    # Weekly / higher TF has fixed factor params 1.7 to 3.0.
    has_factor_30 = any(abs(float(r.get("param", -999)) - 3.0) < 1e-9 for r in results)

    if has_factor_30:
        default = next((r for r in results if abs(float(r.get("param", -999)) - 3.0) < 1e-9), results[-1])
        default_param = 3.0
        default_factor = 3.0
        default_atr = 10
    else:
        default = next((r for r in results if int(r.get("param", -999)) == 20), results[-1])
        default_param = default.get("param", 10)
        default_factor = default.get("factor", 3.0)
        default_atr = default.get("atr", 10)

    fallback = dict(default)
    fallback["valid"] = False
    fallback["bt"] = None
    fallback["atr"] = default_atr
    fallback["factor"] = default_factor
    fallback["factor_now"] = default_factor
    fallback["param"] = 20 if not has_factor_30 else default_param
    fallback["neighbor_avg"] = None
    fallback["plateau_score"] = 1.0
    fallback["stability_score"] = None
    fallback["stability_label"] = "NO BT"
    fallback["roi_winner_param"] = default_param
    fallback["roi_winner_bt"] = None
    fallback["stab_winner_param"] = default_param
    fallback["stab_winner_bt"] = None
    fallback["mode"] = "DEFAULT"
    fallback["changed_by_stability"] = False
    return fallback


def select_best_with_stability(results):
    valid = add_stability_metrics(results)

    if valid:
        stability_winner = select_stability_winner(valid)
        roi_winner = select_roi_winner(valid)
        final_winner = stability_winner if USE_STABILITY_SELECTION else roi_winner

        final_winner["roi_winner_param"] = roi_winner["param"] if roi_winner else None
        final_winner["roi_winner_bt"] = roi_winner["bt"] if roi_winner else None
        final_winner["stab_winner_param"] = stability_winner["param"] if stability_winner else None
        final_winner["stab_winner_bt"] = stability_winner["bt"] if stability_winner else None
        final_winner["mode"] = "STAB" if USE_STABILITY_SELECTION else "ROI"
        final_winner["changed_by_stability"] = (
            roi_winner is not None and stability_winner is not None and roi_winner["param"] != stability_winner["param"]
        )
        return final_winner

    if results:
        return default_no_bt_result(results)

    return None


def optimize_daily(df):
    results = []

    for atr_len in DAILY_ATRS:
        factor_series = daily_factor(df, atr_len)
        st = supertrend_tv(df, factor_series, atr_len)

        if st.empty:
            continue

        bt, used_trades, all_trades, valid = calculate_trade_count_return(
            st, DAILY_PREFERRED_TRADES, DAILY_MINIMUM_TRADES
        )

        factor_clean = factor_series.dropna()
        factor_now = round(float(factor_clean.iloc[-1]), 2) if not factor_clean.empty else np.nan

        results.append({
            "param": int(atr_len),
            "atr": int(atr_len),
            "factor": factor_now,
            "factor_now": factor_now,
            "bt": bt,
            "used_trades": used_trades,
            "all_trades": all_trades,
            "valid": valid,
            "st": st
        })

    return select_best_with_stability(results)


def optimize_weekly(df):
    results = []

    for factor in WEEKLY_FACTORS:
        factor_series = pd.Series(float(factor), index=df.index)
        st = supertrend_tv(df, factor_series, 10)

        if st.empty:
            continue

        bt, used_trades, all_trades, valid = calculate_trade_count_return(
            st, HIGHER_PREFERRED_TRADES, HIGHER_MINIMUM_TRADES
        )

        results.append({
            "param": float(factor),
            "atr": 10,
            "factor": float(factor),
            "factor_now": float(factor),
            "bt": bt,
            "used_trades": used_trades,
            "all_trades": all_trades,
            "valid": valid,
            "st": st
        })

    return select_best_with_stability(results)


def run_logic(df, timeframe):
    if timeframe in ["1w", "1M"]:
        return optimize_weekly(df)
    return optimize_daily(df)


def flip_metrics(st):
    current_trend = st["trend"].iloc[-1]
    flip_index = 0

    for i in range(len(st) - 2, -1, -1):
        if st["trend"].iloc[i] != current_trend:
            flip_index = i + 1
            break

    flip_time = st["time"].iloc[flip_index]
    now = st["time"].iloc[-1]
    days = (now - flip_time).days
    flip_price = st["close"].iloc[flip_index]
    current_price = st["close"].iloc[-1]
    pct = ((current_price - flip_price) / flip_price) * 100.0
    return flip_time.strftime("%Y-%m-%d"), days, round(pct, 2)


# =====================================================
# WORKERS
# =====================================================

class WorkerSignals(QObject):
    result = Signal(list, str, int)
    finished = Signal(str, int)


class CombinedWorker(QRunnable):
    def __init__(self, item, asset_class, timeframe, tab_name, scan_id, force_reoptimize=False):
        super().__init__()
        self.item = item
        self.asset_class = asset_class
        self.timeframe = timeframe
        self.tab_name = tab_name
        self.scan_id = scan_id
        self.force_reoptimize = force_reoptimize
        self.signals = WorkerSignals()

    def run(self):
        try:
            if self.asset_class == "Crypto":
                row = self.run_crypto()
            elif self.asset_class == "Stocks":
                row = self.run_stock()
            else:
                row = self.run_real_asset()

            if row:
                self.signals.result.emit(row, self.tab_name, self.scan_id)

        except Exception:
            traceback.print_exc()

        finally:
            self.signals.finished.emit(self.tab_name, self.scan_id)

    def run_crypto(self):
        name = self.item.get("name", "")
        symbol_raw = self.item.get("symbol", "").lower()

        if "usd" in name.lower() or "usd" in symbol_raw:
            return None

        symbol = symbol_raw.upper()
        route = get_crypto_route(symbol)
        df = fetch_crypto(route, self.timeframe, limit=1000)
        df = drop_unconfirmed_signal_bar(df, self.timeframe)

        if df is None or len(df) < MIN_BARS_CRYPTO:
            return None

        winner, cache_age = run_logic_with_cache(df, self.timeframe, symbol, "Crypto", self.force_reoptimize)
        if winner is None:
            return None

        st = winner["st"]
        trend = "Bullish" if int(st["trend"].iloc[-1]) == 1 else "Bearish"
        flip_date, days, pct = flip_metrics(st)
        market_cap = self.item.get("market_cap", 0) or 0
        market_cap_m = round(market_cap / 1_000_000, 2)
        bt_display = "NA" if winner["bt"] is None else round(winner["bt"], 2)

        return [
            name,
            route["display"],
            round(float(st["close"].iloc[-1]), 6),
            trend,
            flip_date,
            days,
            pct,
            winner["atr"],
            winner["factor"],
            "NA" if winner.get("roi_winner_param") is None else winner.get("roi_winner_param"),
            "NA" if winner.get("stab_winner_param") is None else winner.get("stab_winner_param"),
            "NA" if winner.get("roi_winner_bt") is None else round(winner.get("roi_winner_bt"), 2),
            "NA" if winner.get("stab_winner_bt") is None else round(winner.get("stab_winner_bt"), 2),
            "YES" if winner.get("changed_by_stability") else "NO",
            winner["used_trades"],
            winner["all_trades"],
            "OK" if winner["valid"] else "NO",
            market_cap_m,
            cache_age,
            f"https://www.tradingview.com/chart/?symbol={route['tv_symbol']}"
        ]

    def run_stock(self):
        tv_name = self.item.get("name", "")
        exchange = self.item.get("exchange", "NASDAQ") or "NASDAQ"
        name = self.item.get("description", "") or tv_name
        yf_symbol = clean_yfinance_symbol(tv_name)

        if not yf_symbol or "/" in yf_symbol or len(yf_symbol) > 12:
            return None

        df = fetch_yfinance(yf_symbol)
        if df is None:
            return None

        if self.timeframe == "1w":
            df = to_weekly(df)
            df = drop_unconfirmed_signal_bar(df, self.timeframe)
            if len(df) < MIN_BARS_WEEKLY:
                return None
        else:
            df = drop_unconfirmed_signal_bar(df, self.timeframe)
            if len(df) < MIN_BARS_DAILY:
                return None

        winner, cache_age = run_logic_with_cache(df, self.timeframe, yf_symbol, "Stocks", self.force_reoptimize)
        if winner is None or winner["st"].empty:
            return None

        st = winner["st"]
        trend = "Bullish" if int(st["trend"].iloc[-1]) == 1 else "Bearish"
        flip_date, days, pct = flip_metrics(st)
        market_cap = self.item.get("market_cap_basic", 0) or 0
        market_cap_m = round(float(market_cap) / 1_000_000, 2) if market_cap else 0
        bt_display = "NA" if winner["bt"] is None else round(winner["bt"], 2)
        tv_symbol = clean_tv_symbol(tv_name, exchange)

        return [
            name,
            yf_symbol,
            round(float(st["close"].iloc[-1]), 2),
            trend,
            flip_date,
            days,
            pct,
            winner["atr"],
            winner["factor"],
            "NA" if winner.get("roi_winner_param") is None else winner.get("roi_winner_param"),
            "NA" if winner.get("stab_winner_param") is None else winner.get("stab_winner_param"),
            "NA" if winner.get("roi_winner_bt") is None else round(winner.get("roi_winner_bt"), 2),
            "NA" if winner.get("stab_winner_bt") is None else round(winner.get("stab_winner_bt"), 2),
            "YES" if winner.get("changed_by_stability") else "NO",
            winner["used_trades"],
            winner["all_trades"],
            "OK" if winner["valid"] else "NO",
            market_cap_m,
            cache_age,
            f"https://www.tradingview.com/chart/?symbol={tv_symbol}"
        ]

    def run_real_asset(self):
        symbol = clean_yfinance_symbol(self.item.get("symbol", ""))
        name = self.item.get("name", symbol)

        if not symbol:
            return None

        df = fetch_yfinance(symbol)
        if df is None:
            return None

        if self.timeframe == "1w":
            df = to_weekly(df)
            df = drop_unconfirmed_signal_bar(df, self.timeframe)
            if len(df) < MIN_BARS_WEEKLY:
                return None
        else:
            df = drop_unconfirmed_signal_bar(df, self.timeframe)
            if len(df) < MIN_BARS_DAILY:
                return None

        winner, cache_age = run_logic_with_cache(df, self.timeframe, symbol, "Real Assets", self.force_reoptimize)
        if winner is None or winner["st"].empty:
            return None

        st = winner["st"]
        trend = "Bullish" if int(st["trend"].iloc[-1]) == 1 else "Bearish"
        flip_date, days, pct = flip_metrics(st)
        bt_display = "NA" if winner["bt"] is None else round(winner["bt"], 2)
        tv_symbol = clean_tv_symbol(symbol)

        return [
            name,
            symbol,
            round(float(st["close"].iloc[-1]), 2),
            trend,
            flip_date,
            days,
            pct,
            winner["atr"],
            winner["factor"],
            "NA" if winner.get("roi_winner_param") is None else winner.get("roi_winner_param"),
            "NA" if winner.get("stab_winner_param") is None else winner.get("stab_winner_param"),
            "NA" if winner.get("roi_winner_bt") is None else round(winner.get("roi_winner_bt"), 2),
            "NA" if winner.get("stab_winner_bt") is None else round(winner.get("stab_winner_bt"), 2),
            "YES" if winner.get("changed_by_stability") else "NO",
            winner["used_trades"],
            winner["all_trades"],
            "OK" if winner["valid"] else "NO",
            self.item.get("category", ""),
            cache_age,
            f"https://www.tradingview.com/chart/?symbol={tv_symbol}"
        ]


# =====================================================
# UI
# =====================================================

class NumericItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            return float(self.text()) < float(other.text())
        except Exception:
            return super().__lt__(other)


def format_price_for_display(value):
    try:
        if value is None or pd.isna(value):
            return ""
    except Exception:
        pass

    try:
        x = float(value)
    except Exception:
        return str(value)

    if x == 0:
        return "0"

    if abs(x) < 0.0001:
        return f"{x:.10f}".rstrip("0").rstrip(".")

    if abs(x) < 1:
        return f"{x:.8f}".rstrip("0").rstrip(".")

    if abs(x) < 100:
        return f"{x:.4f}".rstrip("0").rstrip(".")

    return f"{x:.2f}".rstrip("0").rstrip(".")


class ScannerTab(QWidget):
    def __init__(self, timeframe, tab_name, parent_app):
        super().__init__()
        self.timeframe = timeframe
        self.tab_name = tab_name
        self.parent_app = parent_app
        self.dev_view = False
        self.results = []
        self.processed = 0
        self.total_tasks = 0

        layout = QVBoxLayout()
        self.status = QLabel(f"{tab_name}: ready")
        layout.addWidget(self.status)

        self.progress = QProgressBar()
        layout.addWidget(self.progress)

        self.bull_label = QLabel("Bullish: 0")
        self.bear_label = QLabel("Bearish: 0")
        layout.addWidget(self.bull_label)
        layout.addWidget(self.bear_label)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.cellClicked.connect(self.open_link)
        layout.addWidget(self.table)
        self.setLayout(layout)

    def reset(self, total_tasks):
        self.results = []
        self.processed = 0
        self.total_tasks = total_tasks
        self.progress.setValue(0)
        self.progress.show()
        self.status.setText(f"{self.tab_name}: scanning...")
        self.bull_label.setText("Bullish: 0")
        self.bear_label.setText("Bearish: 0")
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def add_result(self, row):
        target = self.parent_app.asset_count_spin.value()

        if len(self.results) >= target:
            self.progress.setValue(100)
            self.progress.hide()
            self.status.setText(f"{self.tab_name}: done, {len(self.results)} results")
            return

        self.results.append(row)
        self.update_table()

        if len(self.results) >= target:
            self.progress.setValue(100)
            self.progress.hide()
            self.status.setText(f"{self.tab_name}: done, {len(self.results)} results")

    def task_done(self):
        self.processed += 1

        if self.total_tasks > 0:
            self.progress.setValue(int(self.processed / self.total_tasks * 100))

        if self.processed >= self.total_tasks:
            self.progress.setValue(100)
            self.progress.hide()
            self.status.setText(f"{self.tab_name}: done, {len(self.results)} results")

    def set_dev_view(self, enabled):
        self.dev_view = bool(enabled)
        self.update_table()

    def update_table(self):
        asset_class = self.parent_app.asset_class_combo.currentText()
        metric_header = "MarketCap (M)" if asset_class in ["Crypto", "Stocks"] else "Category"

        headers = [
            "Asset", "Symbol", "Price", "Trend", "Flip Date", "Days Since Flip",
            "% Since Flip", "Winner ATR", "Winner Factor",
            "ROI Factor/ATR", "STAB Factor/ATR", "ROI BT %", "ROI %",
            "Changed", "Used Trades", "All Trades", "Valid", metric_header, "Cache Age", "Chart"
        ]

        if asset_class in ["Crypto", "Stocks"]:
            sorted_results = sorted(self.results, key=lambda x: x[17] if isinstance(x[17], (int, float)) else 0, reverse=True)
        else:
            sorted_results = list(self.results)

        display = sorted_results
        bull_count = sum(1 for r in display if r[3] == "Bullish")
        bear_count = sum(1 for r in display if r[3] == "Bearish")

        self.bull_label.setText(f"<span style='color:#00e676; font-weight:bold;'>Bullish: {bull_count}</span>")
        self.bear_label.setText(f"<span style='color:#ff5252; font-weight:bold;'>Bearish: {bear_count}</span>")
        self.bull_label.setTextFormat(Qt.RichText)
        self.bear_label.setTextFormat(Qt.RichText)

        self.table.setSortingEnabled(False)
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(display))

        numeric_cols = [2, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15]
        if asset_class in ["Crypto", "Stocks"]:
            numeric_cols.append(17)

        for i, row in enumerate(display):
            for j, val in enumerate(row):
                if j == 19:
                    item = QTableWidgetItem("Open")
                    item.setData(Qt.UserRole, val)
                elif j == 2:
                    item = NumericItem(format_price_for_display(val))
                elif j in numeric_cols:
                    item = NumericItem(str(val))
                else:
                    item = QTableWidgetItem(str(val))

                if j == 3:
                    if val == "Bullish":
                        item.setBackground(Qt.darkGreen)
                        item.setForeground(Qt.white)
                    else:
                        item.setBackground(Qt.darkRed)
                        item.setForeground(Qt.white)

                if j == 18 and str(val).startswith("OLD"):
                    item.setBackground(Qt.darkYellow)
                    item.setForeground(Qt.black)

                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(i, j, item)

        dev_cols = [7, 8, 9, 10, 11, 13, 15, 16, 18]
        for col in dev_cols:
            self.table.setColumnHidden(col, not self.dev_view)

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSortingEnabled(True)

    def open_link(self, row, col):
        if col == 19:
            item = self.table.item(row, col)
            if item:
                url = item.data(Qt.UserRole)
                QDesktopServices.openUrl(QUrl(url))


class App(QMainWindow):
    def __init__(self):
        super().__init__()
        init_factor_cache()
        self.setWindowTitle("Synchromancy Combined Scanner, ROI vs STAB Factor Cache")
        self.resize(1650, 900)

        self.threadpool = QThreadPool()
        self.threadpool.setMaxThreadCount(2)
        self.scan_id = 0

        root = QWidget()
        layout = QVBoxLayout()
        controls = QHBoxLayout()

        controls.addWidget(QLabel("Asset Class:"))
        self.asset_class_combo = QComboBox()
        self.asset_class_combo.addItems(["Crypto", "Stocks", "Real Assets"])
        self.asset_class_combo.currentTextChanged.connect(self.asset_class_changed)
        controls.addWidget(self.asset_class_combo)

        controls.addWidget(QLabel("Nr of Assets:"))
        self.asset_count_spin = QSpinBox()
        controls.addWidget(self.asset_count_spin)

        self.button = QPushButton("Scan Selected Tab")
        self.button.clicked.connect(self.run_scan)
        controls.addWidget(self.button)

        self.reopt_button = QPushButton("Re-optimize Current Market")
        self.reopt_button.clicked.connect(self.run_reoptimize)
        controls.addWidget(self.reopt_button)

        self.export_button = QPushButton("Export Excel")
        self.export_button.clicked.connect(self.export_excel)
        controls.addWidget(self.export_button)

        self.dev_view_cb = QCheckBox("Dev View")
        self.dev_view_cb.setChecked(False)
        self.dev_view_cb.toggled.connect(self.toggle_dev_view)
        controls.addWidget(self.dev_view_cb)

        controls.addStretch()
        layout.addLayout(controls)

        self.tabs = QTabWidget()
        self.weekly_tab = ScannerTab("1w", "Weekly", self)
        self.daily_tab = ScannerTab("1d", "Daily", self)
        self.tabs.addTab(self.weekly_tab, "Weekly")
        self.tabs.addTab(self.daily_tab, "Daily")
        layout.addWidget(self.tabs)

        root.setLayout(layout)
        self.setCentralWidget(root)
        self.asset_class_changed("Crypto")

    def toggle_dev_view(self, checked):
        self.weekly_tab.set_dev_view(checked)
        self.daily_tab.set_dev_view(checked)

    def save_current_count(self):
        asset_class = self.asset_class_combo.currentText()
        if asset_class in ASSET_CLASS_SETTINGS:
            ASSET_CLASS_SETTINGS[asset_class]["count"] = self.asset_count_spin.value()

    def asset_class_changed(self, asset_class):
        settings = ASSET_CLASS_SETTINGS[asset_class]
        absolute_max = settings["absolute_max"]

        self.asset_count_spin.blockSignals(True)
        self.asset_count_spin.setRange(1, absolute_max)
        self.asset_count_spin.setValue(settings["count"])
        self.asset_count_spin.blockSignals(False)

    def get_active_tab(self):
        if self.tabs.currentIndex() == 0:
            return self.weekly_tab
        return self.daily_tab

    def build_universe(self, asset_class, count):
        count = max(1, min(count, ASSET_CLASS_SETTINGS[asset_class]["absolute_max"]))

        if asset_class == "Crypto":
            # Restore old crypto behavior: if user asks for 50 final rows,
            # scan many more CoinGecko candidates because many get rejected.
            candidate_count = min(count * 20, ASSET_CLASS_SETTINGS[asset_class]["absolute_max"])
            return get_crypto_universe(1, candidate_count)

        if asset_class == "Stocks":
            # Light over-fetch for delisted/bad Yahoo symbols without making stock scans too heavy.
            candidate_count = min(count * 3, ASSET_CLASS_SETTINGS[asset_class]["absolute_max"])
            return get_stock_universe(1, candidate_count)

        return get_real_asset_universe(1, count)

    def run_scan(self):
        self.run_scan_internal(force_reoptimize=False)

    def run_reoptimize(self):
        self.run_scan_internal(force_reoptimize=True)

    def run_scan_internal(self, force_reoptimize=False):
        self.save_current_count()
        self.scan_id += 1
        current_scan = self.scan_id

        asset_class = self.asset_class_combo.currentText()
        count = self.asset_count_spin.value()

        universe = self.build_universe(asset_class, count)
        tab = self.get_active_tab()
        tab.reset(len(universe))

        if force_reoptimize:
            tab.status.setText(f"{tab.tab_name}: re-optimizing {count} {asset_class} results from {len(universe)} candidates...")
        else:
            tab.status.setText(f"{tab.tab_name}: scanning for {count} {asset_class} results from {len(universe)} candidates...")

        if not universe:
            tab.task_done()
            tab.status.setText(f"{tab.tab_name}: no assets found")
            return

        for item in universe:
            worker = CombinedWorker(item, asset_class, tab.timeframe, tab.tab_name, current_scan, force_reoptimize=force_reoptimize)
            worker.signals.result.connect(self.add_result)
            worker.signals.finished.connect(self.task_done)
            self.threadpool.start(worker)

    def add_result(self, row, tab_name, scan_id):
        if scan_id != self.scan_id:
            return

        if tab_name == "Weekly":
            self.weekly_tab.add_result(row)
        elif tab_name == "Daily":
            self.daily_tab.add_result(row)

    def task_done(self, tab_name, scan_id):
        if scan_id != self.scan_id:
            return

        if tab_name == "Weekly":
            self.weekly_tab.task_done()
        elif tab_name == "Daily":
            self.daily_tab.task_done()

    def export_excel(self):
        asset_class = self.asset_class_combo.currentText()
        tab = self.get_active_tab()
        data = tab.results
        metric_header = "MarketCap (M)" if asset_class in ["Crypto", "Stocks"] else "Category"
        filename = f"{asset_class} Scan {tab.tab_name}.xlsx".replace(" ", "_")

        headers = [
            "Asset", "Symbol", "Price", "Trend", "Flip Date",
            "Days Since Flip", "% Since Flip", "Winner ATR",
            "Winner Factor", "ROI Factor/ATR", "STAB Factor/ATR",
            "ROI BT %", "STAB BT %", "Changed", "Used Trades",
            "All Trades", "Valid", metric_header, "Cache Age", "Chart"
        ]

        df = pd.DataFrame(data, columns=headers)

        from openpyxl.styles import Font, PatternFill, Alignment

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Scanner")
            ws = writer.sheets["Scanner"]

            header_fill = PatternFill(fill_type="solid", fgColor="1565C0")
            bull_fill = PatternFill(fill_type="solid", fgColor="006400")
            bear_fill = PatternFill(fill_type="solid", fgColor="8B0000")
            white_bold = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center")

            for row in range(2, ws.max_row + 1):
                trend_cell = ws[f"D{row}"]
                if trend_cell.value == "Bullish":
                    trend_cell.fill = bull_fill
                    trend_cell.font = white_bold
                elif trend_cell.value == "Bearish":
                    trend_cell.fill = bear_fill
                    trend_cell.font = white_bold

                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for row in range(2, ws.max_row + 1):
                link_cell = ws[f"T{row}"]
                url = link_cell.value
                if url:
                    link_cell.hyperlink = url
                    link_cell.value = "Open Chart"
                    link_cell.font = Font(color="00B0F0", underline="single", bold=True)

        print(f"Exported {filename}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet("""
    QWidget {
        background-color: #121212;
        color: #e0e0e0;
        font-size: 12px;
    }
    QTableWidget {
        background-color: #1e1e1e;
        gridline-color: #333333;
    }
    QHeaderView::section {
        background-color: #2c2c2c;
        color: white;
        padding: 4px;
    }
    QPushButton {
        background-color: #1565C0;
        color: white;
        padding: 7px;
        font-weight: bold;
    }
    QPushButton:hover {
        background-color: #1E88E5;
    }
    QProgressBar {
        background-color: #2c2c2c;
        color: white;
        text-align: center;
    }
    QProgressBar::chunk {
        background-color: #1E88E5;
    }
    QTabWidget::pane {
        border: 1px solid #333333;
    }
    QTabBar::tab {
        background: #2c2c2c;
        color: white;
        padding: 8px;
    }
    QTabBar::tab:selected {
        background: #1565C0;
    }
    QComboBox, QSpinBox {
        background-color: #1e1e1e;
        color: white;
        padding: 4px;
        border: 1px solid #333333;
    }
    """)

    window = App()
    window.show()
    sys.exit(app.exec())
